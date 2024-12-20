from django.shortcuts import render, redirect, get_object_or_404
from .models import Event, EventType, Province, District, Municipality, Participant, Project, Ethnicity, Occupation, ProjectFile
from .forms import EventForm, ParticipantForm, ProjectForm, EventTypeForm, ProjectFileForm
import openpyxl
from openpyxl import Workbook
from django.contrib import messages
from django.http import HttpResponse, Http404, JsonResponse
from django.db.models import Count
from django.template.loader import render_to_string
from django.urls import reverse


def load_districts(request):
    province_id = request.GET.get('province')
    districts = District.objects.filter(province_id=province_id).order_by('name')
    # return JsonResponse(list(districts.values('id', 'name')), safe=False)
    html = render_to_string('partials/district_options.html', {'districts': districts})
    return HttpResponse(html)

def load_municipalities(request):
    district_id = request.GET.get('district')
    municipalities = Municipality.objects.filter(district_id=district_id).order_by('name')
    # return JsonResponse(list(municipalities.values('id', 'name')), safe=False)
    html = render_to_string('partials/municipality_options.html', {'municipalities': municipalities})
    return HttpResponse(html)


def homepage(request):
    return render(request, 'homepage.html')

def event_list(request):
    events = Event.objects.all()
    return render(request, 'events/event_list.html', {'events': events})

# def event_create(request):
#     if request.method == 'POST':
#         form = EventForm(request.POST)
#         if form.is_valid():
#             form.save()
#             messages.success(request, "Event created successfully.")
#             return redirect('event_list')
#     else:
#         form = EventForm()
#     return render(request, 'events/event_form.html', {'form': form})

# def event_edit(request, pk):
#     event = Event.objects.get(pk=pk)
#     if request.method == 'POST':
#         form = EventForm(request.POST, instance=event)
#         if form.is_valid():
#             form.save()
#             messages.success(request, "Event updated successfully.")
#             return redirect('event_list')
#     else:
#         form = EventForm(instance=event)
#     return render(request, 'events/event_form.html', {'form': form})
# def event_form(request, pk=None):
#     if pk:
#         event = get_object_or_404(Event, pk=pk)
#     else:
#         event = None

#     if request.method == 'POST':
#         form = EventForm(request.POST, instance=event)
#         if form.is_valid():
#             form.save()
#             return redirect('event_list')
#     else:
#         form = EventForm(instance=event)

#     return render(request, 'events/event_form.html', {'form': form})

def event_create(request, project_id=None):
    """
    View to create a new event. Optionally links it to a project if `project_id` is provided.
    """
    project = None
    if project_id:
        project = get_object_or_404(Project, id=project_id)

    if request.method == "POST":
        form = EventForm(request.POST)
        if form.is_valid():
            event = form.save(commit=False)
            if project:
                event.project = project
            event.save()
            return redirect(reverse('project_detail', kwargs={'pk': project.id}) if project else reverse('event_list'))
    else:
        form = EventForm()

    return render(request, 'events/event_form.html', {'form': form, 'project': project})

def event_edit(request, project_id, pk):
    """
    View to edit an existing event.
    """
    event = get_object_or_404(Event, pk=pk)
    if request.method == "POST":
        form = EventForm(request.POST, instance=event)
        if form.is_valid():
            form.save()
            if event.project:
                return redirect(reverse('project_detail', kwargs={'pk': event.project.id}))
            return redirect(reverse('event_list'))
    else:
        form = EventForm(instance=event)

    return render(request, 'events/event_form.html', {'form': form, 'event': event})


def event_delete(request, pk):
    """
    View to delete an event.
    """
    event = get_object_or_404(Event, pk=pk)
    project_id = event.project.id if event.project else None
    event.delete()
    if project_id:
        return redirect(reverse('project_detail', kwargs={'pk': project_id}))
    return redirect(reverse('event_list'))


def event_detail(request, pk):
    event = get_object_or_404(Event, pk=pk)

    # Get participants for the event
    participants = Participant.objects.filter(event=event)

    # Total number of participants
    total_participants = participants.count()

    # Ethnicity counts
    ethnicity_counts = participants.values('ethnicity__name').annotate(count=Count('id'))
    
    gender_counts = participants.values('gender__name').annotate(count=Count('id'))

    # Certified participants count
    certified_count = participants.filter(certification=True).count()

    context = {
        'event': event,
        'participants': participants,
        'total_participants': total_participants,
        'ethnicity_counts': ethnicity_counts,
        'certified_count': certified_count,
        'gender_counts': gender_counts,
    }
    return render(request, 'events/event_detail.html', context)



# List all participants for a given event
def participant_list(request, event_id):
    event = Event.objects.get(id=event_id)
    participants = Participant.objects.filter(event=event)
    return render(request, 'events/participant_list.html', {'participants': participants, 'event': event})

# Create a new participant for an event
def participant_create(request, event_id):
    event = Event.objects.get(id=event_id)
    if request.method == 'POST':
        form = ParticipantForm(request.POST)
        if form.is_valid():
            participant = form.save(commit=False)
            participant.event = event
            participant.save()
            messages.success(request, "Participant added successfully.")
            return redirect('event_detail', pk=event.id)
    else:
        form = ParticipantForm()
    return render(request, 'events/participant_form.html', {'form': form, 'event': event})

# Edit an existing participant
def participant_edit(request, participant_id):
    participant = Participant.objects.get(id=participant_id)
    if request.method == 'POST':
        form = ParticipantForm(request.POST, instance=participant)
        if form.is_valid():
            form.save()
            messages.success(request, "Participant updated successfully.")
            return redirect('event_detail', pk=participant.event.id)
    else:
        form = ParticipantForm(instance=participant)
    return render(request, 'events/participant_form.html', {'form': form, 'event': participant.event})

def participant_detail(request, pk):
    """
    View to display details of a single participant.
    """
    participant = get_object_or_404(Participant, pk=pk)
    return render(request, 'events/participant_detail.html', {'participant': participant})

# Delete a participant
def participant_delete(request, participant_id):
    participant = Participant.objects.get(id=participant_id)
    participant.delete()
    messages.success(request, "Participant deleted successfully.")
    return redirect('participant_list', event_id=participant.event.id)

# Import participants from an Excel file
def import_participants(request, event_id):
    if request.method == 'POST' and request.FILES['file']:
        excel_file = request.FILES['file']
        wb = openpyxl.load_workbook(excel_file)
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
            title, first_name, last_name, email, phone_number, address, gender, ethnicity, disability, occupation, organization, designation, certification = row
            # Find or create ethnicity and occupation
            ethnicity_obj, created = Ethnicity.objects.get_or_create(name=ethnicity)
            occupation_obj, created = Occupation.objects.get_or_create(name=occupation)
            # Create participant
            Participant.objects.create(
                event_id=event_id,
                first_name=first_name,
                last_name=last_name,
                ethnicity=ethnicity_obj,
                occupation=occupation_obj,
                email=email,
                phone_number=phone_number,
                address=address
            )
        messages.success(request, "Participants imported successfully.")
        return redirect('participant_list', event_id=event_id)
    return render(request, 'events/import_participants.html', {'event_id': event_id})


def participant_template(request):
    # Create a workbook and add a worksheet
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Participants Template"

    # Define headers
    headers = [
        "First Name", "Last Name", "Email", "Phone Number", 
        "Ethnicity", "Occupation"
    ]
    worksheet.append(headers)

    # Set column widths for better readability
    for col, header in enumerate(headers, start=1):
        worksheet.cell(row=1, column=col).value = header
        worksheet.column_dimensions[chr(64 + col)].width = 20

    # Set the response headers to serve the Excel file
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="participant_template.xlsx"'

    # Save workbook to response
    workbook.save(response)
    return response

def export_events(request):
    # Create a new Excel workbook and sheet
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Events"

    # Define the headers
    headers = [
        "Name", "Start Date", "End Date",
        "Province", "District", "Municipality", "Ward No",
        "Venue", "Latitude", "Longitude", "Partner Name"
    ]
    worksheet.append(headers)

    # Populate the worksheet with event data
    events = Event.objects.all()
    for event in events:
        worksheet.append([
            event.name,
            event.start_date,
            event.end_date,
            event.province.name if event.province else "",
            event.district.name if event.district else "",
            event.municipality.name if event.municipality else "",
            event.ward_no,
            event.venue,
            event.latitude,
            event.longitude,
            event.organizing_partners,
            event.funding_partners
        ])

    # Prepare the response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="events.xlsx"'

    # Save workbook to the response
    workbook.save(response)
    return response




# PROJECTS
# List projects
def project_list(request):
    projects = Project.objects.all()
    return render(request, 'project/project_list.html', {'projects': projects})

# Create a project
def project_create(request):
    if request.method == 'POST':
        form = ProjectForm(request.POST)
        if form.is_valid():
            project = form.save()
            messages.success(request, "Project created successfully.")
            return redirect('project_detail', pk=project.id)  # Redirect to the project detail page
        else:
            messages.error(request, "There was an error creating the project. Please check the form.")
    else:
        form = ProjectForm()
    return render(request, 'project/project_form.html', {'form': form, 'title': 'Add Project'})

# Update a project
def project_edit(request, pk):
    project = get_object_or_404(Project, pk=pk)
    if request.method == 'POST':
        form = ProjectForm(request.POST, instance=project)
        if form.is_valid():
            form.save()
            return redirect('project_detail', pk=pk)
    else:
        form = ProjectForm(instance=project)
    return render(request, 'project/project_form.html', {'form': form, 'title': 'Edit Project'})

def project_detail(request, pk):
    project = get_object_or_404(Project, pk=pk)
    events = Event.objects.filter(project=project).order_by('start_date')  # Assuming an Event model has a `project` field
    return render(request, 'project/project_detail.html', {'project': project, 'events': events})

# Delete a project
def project_delete(request, pk):
    project = get_object_or_404(Project, pk=pk)
    if request.method == 'POST':
        project.delete()
        return JsonResponse({'status': 'success'})
    return render(request, 'project/project_confirm_delete.html', {'project': project})

def add_project_file(request, project_id):
    """
    View to add a new project file to a specific project.
    """
    project = get_object_or_404(Project, id=project_id)

    if request.method == 'POST':
        form = ProjectFileForm(request.POST, request.FILES)
        # breakpoint()
        if form.is_valid():
            # Associate the file with the project
            project_file = form.save(commit=False)
            project_file.project = project
            project_file.save()
            # form.save()
            # breakpoint()
            messages.success(request, 'Project file has been uploaded successfully!')
            return redirect('project_detail', pk=project.id)
        else:
            messages.error(request, 'There was an error with your file upload.')

    else:
        form = ProjectFileForm()

    return render(request, 'project/add_project_file.html', {'form': form, 'project': project})

def delete_project_file(request, file_id):
    """
    View to delete a specific project file.
    """
    file = get_object_or_404(ProjectFile, id=file_id)
    project_id = file.project.id  # Save project ID to redirect after deletion
    file.delete()
    messages.success(request, 'File deleted successfully.')
    return redirect('project_detail', pk=project_id)


# EVENT TYPES

# List event types
def eventtype_list(request):
    eventtypes = EventType.objects.all()
    return render(request, 'eventtype/eventtype_list.html', {'eventtypes': eventtypes})

# Create event type
def eventtype_create(request):
    if request.method == 'POST':
        form = EventTypeForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('eventtype_list')
    else:
        form = EventTypeForm()
    return render(request, 'eventtype/eventtype_form.html', {'form': form, 'title': 'Add Event Type'})

# Update event type
def eventtype_edit(request, pk):
    eventtype = get_object_or_404(EventType, pk=pk)
    if request.method == 'POST':
        form = EventTypeForm(request.POST, instance=eventtype)
        if form.is_valid():
            form.save()
            return redirect('eventtype_list')
    else:
        form = EventTypeForm(instance=eventtype)
    return render(request, 'eventtype/eventtype_form.html', {'form': form, 'title': 'Edit Event Type'})

# Delete event type
def eventtype_delete(request, pk):
    eventtype = get_object_or_404(EventType, pk=pk)
    if request.method == 'POST':
        eventtype.delete()
        return redirect('eventtype_list')
    return render(request, 'eventtype/eventtype_confirm_delete.html', {'eventtype': eventtype})